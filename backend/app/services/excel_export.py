"""
Excel export service — Milestone 24.2

Produces a .xlsx workbook from a list of Business ORM rows with:
  - Sheet 1: Leads          — all columns, formatted table
  - Sheet 2: Opportunities  — only rows with an opportunity reason
  - Sheet 3: Summary        — export metadata and aggregate stats

Unicode (Cyrillic, Greek, Turkish, emoji, …) is preserved natively because
openpyxl stores all text as UTF-8 internally and xlsx is an XML format.
"""

from __future__ import annotations

import io
import statistics
from datetime import datetime, timezone
from decimal import Decimal
from typing import Any, Sequence

from openpyxl import Workbook
from openpyxl.styles import (
    Alignment,
    Border,
    Font,
    GradientFill,
    PatternFill,
    Side,
)
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet

from app.models.business import Business

# ─── Palette ─────────────────────────────────────────────────────────────────

_HDR_BG     = "1E40AF"  # blue-800
_HDR_FG     = "FFFFFF"
_ROW_ALT    = "F0F4FF"  # very light blue
_ROW_NORM   = "FFFFFF"

# AI-priority row highlight fills
_FILL_HOT   = PatternFill("solid", fgColor="FEE2E2")   # red-100
_FILL_WARM  = PatternFill("solid", fgColor="FFEDD5")   # orange-100
_FILL_QUAL  = PatternFill("solid", fgColor="DBEAFE")   # blue-100
_FILL_COLD  = PatternFill("solid", fgColor="F1F5F9")   # slate-100
_FILL_ALT   = PatternFill("solid", fgColor=_ROW_ALT)
_FILL_WHITE = PatternFill("solid", fgColor=_ROW_NORM)

_THIN_BORDER = Border(
    left=Side(style="thin", color="E2E8F0"),
    right=Side(style="thin", color="E2E8F0"),
    top=Side(style="thin", color="E2E8F0"),
    bottom=Side(style="thin", color="E2E8F0"),
)

# ─── Column definitions ───────────────────────────────────────────────────────
# (header_label, attribute_name, type_hint, min_width)
# type_hint: "text" | "number" | "phone" | "url" | "date" | "pct"

LEADS_COLUMNS: list[tuple[str, str, str, int]] = [
    ("ID",              "id",                   "number", 6),
    ("Name",            "name",                 "text",   28),
    ("Phone",           "phone",                "phone",  16),
    ("Email",           "email",                "text",   28),
    ("Category",        "category",             "text",   20),
    ("City",            "city",                 "text",   16),
    ("Country",         "country",              "text",   16),
    ("Website Status",  "website_status",       "text",   16),
    ("Website",         "website",              "url",    30),
    ("CMS / Platform",  "website_platform",     "text",   16),
    ("Health Score",    "website_health_score", "number", 13),
    ("Redesign Score",  "website_redesign_score","number",14),
    ("AI Score",        "ai_score",             "number", 10),
    ("Est. Value (£)",  "ai_project_value",     "number", 14),
    ("Close %",         "ai_close_prob",        "pct",    9),
    ("AI Priority",     "ai_priority",          "text",   12),
    ("Pipeline",        "contact_status",       "text",   14),
    ("Rating",          "rating",               "number", 8),
    ("Reviews",         "review_count",         "number", 9),
    ("Facebook",        "facebook_url",         "url",    30),
    ("Instagram",       "instagram_url",        "url",    30),
    ("LinkedIn",        "linkedin_url",         "url",    30),
    ("Google Maps",     "google_maps_url",      "url",    30),
    ("Address",         "address",              "text",   30),
    ("Opportunity",     "opportunity_reason",   "text",   35),
    ("Notes",           "notes",                "text",   35),
    ("Added",           "created_at",           "date",   14),
]


# ─── Helpers ──────────────────────────────────────────────────────────────────

def _v(biz: Business, attr: str) -> Any:
    """Return attribute value, coercing Decimal to float."""
    val = getattr(biz, attr, None)
    if isinstance(val, Decimal):
        return float(val)
    return val


def _write_header(ws: Worksheet, columns: list[tuple[str, str, str, int]]) -> None:
    hdr_font  = Font(bold=True, color=_HDR_FG, size=10)
    hdr_fill  = PatternFill("solid", fgColor=_HDR_BG)
    hdr_align = Alignment(horizontal="center", vertical="center", wrap_text=False)

    for col_idx, (label, _, _, _) in enumerate(columns, start=1):
        cell = ws.cell(row=1, column=col_idx, value=label)
        cell.font  = hdr_fill and hdr_font
        cell.fill  = hdr_fill
        cell.alignment = hdr_align
        cell.border = _THIN_BORDER


def _priority_fill(priority: str | None, alt: bool) -> PatternFill | None:
    if priority == "HOT":
        return _FILL_HOT
    if priority == "WARM":
        return _FILL_WARM
    if priority == "QUALIFIED":
        return _FILL_QUAL
    if priority == "COLD":
        return _FILL_COLD
    return _FILL_ALT if alt else _FILL_WHITE


def _write_row(
    ws: Worksheet,
    row_idx: int,
    biz: Business,
    columns: list[tuple[str, str, str, int]],
    alt: bool,
) -> None:
    priority = getattr(biz, "ai_priority", None)
    row_fill = _priority_fill(priority, alt)

    for col_idx, (_, attr, type_hint, _) in enumerate(columns, start=1):
        val = _v(biz, attr)
        cell = ws.cell(row=row_idx, column=col_idx)

        if type_hint == "phone":
            # Store phone as text to preserve leading zeros / + prefix
            if val is not None:
                cell.value = str(val)
                cell.number_format = "@"
            else:
                cell.value = None

        elif type_hint == "url":
            if val:
                url = str(val)
                if not url.startswith(("http://", "https://")):
                    url = "https://" + url
                cell.value = url
                cell.hyperlink = url
                cell.font = Font(color="1D4ED8", underline="single", size=9)
            else:
                cell.value = None

        elif type_hint == "date":
            if isinstance(val, datetime):
                cell.value = val.replace(tzinfo=None) if val.tzinfo else val
                cell.number_format = "DD/MM/YYYY"
            else:
                cell.value = val

        elif type_hint == "pct":
            cell.value = val  # stored as int (e.g. 75 = 75%)

        elif type_hint == "number":
            cell.value = val  # already float or int or None

        else:  # text
            cell.value = str(val) if val is not None else None

        if row_fill and type_hint != "url":
            cell.fill = row_fill
        elif row_fill and type_hint == "url" and not val:
            cell.fill = row_fill

        cell.alignment = Alignment(vertical="top", wrap_text=(type_hint == "text"))
        cell.border = _THIN_BORDER


def _auto_width(ws: Worksheet, columns: list[tuple[str, str, str, int]]) -> None:
    for col_idx, (label, _, _, min_w) in enumerate(columns, start=1):
        col_letter = get_column_letter(col_idx)
        # Sample up to 200 rows for width estimation
        max_len = len(label) + 2
        for row_idx in range(2, min(ws.max_row + 1, 202)):
            cell = ws.cell(row=row_idx, column=col_idx)
            if cell.value is not None:
                max_len = max(max_len, min(len(str(cell.value)), 60))
        ws.column_dimensions[col_letter].width = max(min_w, max_len + 2)


# ─── Leads sheet ─────────────────────────────────────────────────────────────

def _build_leads_sheet(ws: Worksheet, businesses: Sequence[Business]) -> None:
    ws.title = "Leads"
    ws.freeze_panes = "A2"
    ws.sheet_view.showGridLines = True
    ws.row_dimensions[1].height = 22

    _write_header(ws, LEADS_COLUMNS)

    for i, biz in enumerate(businesses):
        row_idx = i + 2
        _write_row(ws, row_idx, biz, LEADS_COLUMNS, alt=(i % 2 == 1))

    if businesses:
        from openpyxl.worksheet.filters import AutoFilter
        last_col = get_column_letter(len(LEADS_COLUMNS))
        last_row = len(businesses) + 1
        ws.auto_filter.ref = f"A1:{last_col}{last_row}"

    _auto_width(ws, LEADS_COLUMNS)


# ─── Opportunities sheet ──────────────────────────────────────────────────────

OPP_COLUMNS: list[tuple[str, str, str, int]] = [
    ("Name",           "name",               "text",   28),
    ("Phone",          "phone",              "phone",  16),
    ("Email",          "email",              "text",   28),
    ("City",           "city",               "text",   16),
    ("Category",       "category",           "text",   20),
    ("Website Status", "website_status",     "text",   16),
    ("Opportunity",    "opportunity_reason", "text",   40),
    ("AI Score",       "ai_score",           "number", 10),
    ("Est. Value (£)", "ai_project_value",   "number", 14),
    ("Close %",        "ai_close_prob",      "pct",    9),
    ("AI Priority",    "ai_priority",        "text",   12),
    ("Website",        "website",            "url",    30),
    ("Google Maps",    "google_maps_url",    "url",    30),
    ("Added",          "created_at",         "date",   14),
]


def _build_opportunities_sheet(ws: Worksheet, businesses: Sequence[Business]) -> None:
    ws.title = "Opportunities"
    ws.freeze_panes = "A2"

    opps = [b for b in businesses if b.opportunity_reason]
    _write_header(ws, OPP_COLUMNS)

    for i, biz in enumerate(opps):
        _write_row(ws, i + 2, biz, OPP_COLUMNS, alt=(i % 2 == 1))

    if opps:
        last_col = get_column_letter(len(OPP_COLUMNS))
        ws.auto_filter.ref = f"A1:{last_col}{len(opps) + 1}"

    _auto_width(ws, OPP_COLUMNS)


# ─── Summary sheet ────────────────────────────────────────────────────────────

def _build_summary_sheet(ws: Worksheet, businesses: Sequence[Business]) -> None:
    ws.title = "Summary"
    ws.column_dimensions["A"].width = 30
    ws.column_dimensions["B"].width = 22

    title_font  = Font(bold=True, size=14, color=_HDR_BG)
    label_font  = Font(bold=True, size=10)
    value_font  = Font(size=10)
    section_font = Font(bold=True, size=11, color=_HDR_BG)
    hdr_fill    = PatternFill("solid", fgColor="DBEAFE")

    def row(r: int, label: str, value: Any) -> None:
        a = ws.cell(row=r, column=1, value=label)
        b = ws.cell(row=r, column=2, value=value)
        a.font = label_font
        b.font = value_font
        b.alignment = Alignment(horizontal="left")

    def section(r: int, title: str) -> None:
        c = ws.cell(row=r, column=1, value=title)
        c.font = section_font
        c.fill = hdr_fill
        ws.cell(row=r, column=2).fill = hdr_fill
        ws.merge_cells(f"A{r}:B{r}")

    # Title
    t = ws.cell(row=1, column=1, value="LeadScrapper Export Summary")
    t.font = title_font
    ws.merge_cells("A1:B1")
    ws.row_dimensions[1].height = 24

    now = datetime.now(tz=timezone.utc)
    row(2, "Export Date", now.strftime("%d %b %Y %H:%M UTC"))
    row(3, "Total Leads", len(businesses))

    # ── Opportunity stats ──
    section(5, "Opportunities")
    opps = [b for b in businesses if b.opportunity_reason]
    no_web  = sum(1 for b in businesses if b.website_status == "NO_WEBSITE")
    fb_only = sum(1 for b in businesses if b.website_status == "FACEBOOK_ONLY")
    free    = sum(1 for b in businesses if b.website_status == "FREE_BUILDER")
    row(6,  "Total Opportunities",  len(opps))
    row(7,  "No Website",           no_web)
    row(8,  "Facebook Only",        fb_only)
    row(9,  "Free Builder",         free)

    # ── AI stats ──
    section(11, "AI Sales Intelligence")
    ai_scored   = [b.ai_score for b in businesses if b.ai_score is not None]
    ai_values   = [b.ai_project_value for b in businesses if b.ai_project_value is not None]
    hot   = sum(1 for b in businesses if b.ai_priority == "HOT")
    warm  = sum(1 for b in businesses if b.ai_priority == "WARM")
    qual  = sum(1 for b in businesses if b.ai_priority == "QUALIFIED")
    cold  = sum(1 for b in businesses if b.ai_priority == "COLD")
    row(12, "AI Analysed",         len(ai_scored))
    row(13, "Avg AI Score",        round(statistics.mean(ai_scored), 1) if ai_scored else "—")
    row(14, "Total Est. Value (£)", sum(ai_values) if ai_values else 0)
    row(15, "HOT Leads",           hot)
    row(16, "WARM Leads",          warm)
    row(17, "QUALIFIED Leads",     qual)
    row(18, "COLD Leads",          cold)

    # ── Website stats ──
    section(20, "Website Statistics")
    has_site  = sum(1 for b in businesses if b.website_status == "HAS_WEBSITE")
    broken    = sum(1 for b in businesses if b.website_status == "BROKEN_WEBSITE")
    analysed  = sum(1 for b in businesses if b.website_health_score is not None)
    health_s  = [b.website_health_score for b in businesses if b.website_health_score is not None]
    row(21, "Has Website",         has_site)
    row(22, "Broken Website",      broken)
    row(23, "Website Analysed",    analysed)
    row(24, "Avg Health Score",    round(statistics.mean(health_s), 1) if health_s else "—")
    has_email = sum(1 for b in businesses if b.email)
    has_phone = sum(1 for b in businesses if b.phone)
    row(25, "Has Email",           has_email)
    row(26, "Has Phone",           has_phone)

    # ── Pipeline stats ──
    section(28, "Pipeline")
    pipeline_counts: dict[str, int] = {}
    for b in businesses:
        pipeline_counts[b.contact_status] = pipeline_counts.get(b.contact_status, 0) + 1
    r = 29
    for stage in ["NEW", "CALLED", "NO_ANSWER", "INTERESTED", "FOLLOW_UP", "PROPOSAL_SENT", "WON", "LOST"]:
        row(r, stage.replace("_", " ").title(), pipeline_counts.get(stage, 0))
        r += 1


# ─── Public API ───────────────────────────────────────────────────────────────

def build_excel_export(businesses: Sequence[Business]) -> bytes:
    """Return raw .xlsx bytes for the given list of Business rows."""
    wb = Workbook()

    # openpyxl creates a default "Sheet" — use it for Leads
    ws_leads = wb.active
    _build_leads_sheet(ws_leads, businesses)

    ws_opps = wb.create_sheet()
    _build_opportunities_sheet(ws_opps, businesses)

    ws_summary = wb.create_sheet()
    _build_summary_sheet(ws_summary, businesses)

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.read()
